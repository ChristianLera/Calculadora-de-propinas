"""
================================================================================
CALCULADORA DE PROPINAS - VERSIÓN SIMPLE PERO PROFESIONAL
================================================================================
Un programa sencillo pero completo para calcular propinas.
Incluye: historial, estadísticas, exportación a Excel y múltiples monedas.
================================================================================
"""

import json
import os
import csv
from datetime import datetime
from typing import List, Dict, Optional
from enum import Enum
from dataclasses import dataclass, asdict
from decimal import Decimal, ROUND_HALF_UP


# =================================================================================
# CONFIGURACIÓN BÁSICA
# =================================================================================

class CalidadServicio(Enum):
    """Calidades de servicio con sus porcentajes recomendados"""
    PESIMO = (0, "Pésimo (0%) - Servicio horrible")
    MALO = (5, "Malo (5%) - Servicio deficiente")
    REGULAR = (10, "Regular (10%) - Servicio aceptable")
    BUENO = (15, "Bueno (15%) - Servicio satisfactorio")
    MUY_BUENO = (18, "Muy Bueno (18%) - Servicio destacable")
    EXCELENTE = (20, "Excelente (20%) - Servicio sobresaliente")
    EXTRAORDINARIO = (25, "Extraordinario (25%) - Servio excepcional")
    
    def __init__(self, porcentaje: int, descripcion: str):
        self.porcentaje = porcentaje
        self.descripcion = descripcion


class Moneda(Enum):
    """Monedas soportadas por la calculadora"""
    USD = ("USD", "$")      # Dólar Americano
    EUR = ("EUR", "€")      # Euro
    MXN = ("MXN", "$")      # Peso Mexicano
    GBP = ("GBP", "£")      # Libra Esterlina
    COP = ("COP", "$")      # Peso Colombiano
    ARS = ("ARS", "$")      # Peso Argentino
    
    def __init__(self, codigo: str, simbolo: str):
        self.codigo = codigo
        self.simbolo = simbolo


# =================================================================================
# MODELO DE DATOS PARA CADA CÁLCULO
# =================================================================================

@dataclass
class CalculoPropina:
    """Estructura que guarda un cálculo completo de propina"""
    fecha: str                      # Fecha y hora del cálculo
    monto_original: float           # Monto original de la cuenta
    moneda: str                     # Código de moneda (USD, EUR, etc.)
    porcentaje_propina: float       # Porcentaje de propina aplicado
    calidad_servicio: str           # Descripción de la calidad seleccionada
    propina: float                  # Monto de la propina calculada
    total: float                    # Monto total (cuenta + propina)
    num_personas: int               # Número de personas (1 si no se divide)
    total_por_persona: float        # Total que paga cada persona
    
    def to_dict(self) -> Dict:
        """Convierte el objeto a diccionario para guardar en JSON"""
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'CalculoPropina':
        """Crea un objeto desde un diccionario (para cargar del JSON)"""
        return cls(**data)


# =================================================================================
# GESTOR DEL HISTORIAL (GUARDA Y CARGA CÁLCULOS)
# =================================================================================

class GestorHistorial:
    """Maneja el guardado y carga del historial de cálculos"""
    
    def __init__(self):
        """Inicializa el gestor y carga el historial existente"""
        self.archivo_historial = "historial_propinas.json"
        self.historial: List[CalculoPropina] = []
        self.cargar_historial()
    
    def cargar_historial(self):
        """Carga los cálculos guardados desde el archivo JSON"""
        try:
            if os.path.exists(self.archivo_historial):
                with open(self.archivo_historial, 'r', encoding='utf-8') as f:
                    datos = json.load(f)
                    # Convierte cada diccionario del JSON en un objeto CalculoPropina
                    self.historial = [CalculoPropina.from_dict(item) for item in datos]
        except Exception as e:
            # Si hay error al cargar, empieza con historial vacío
            print(f"⚠️ No se pudo cargar el historial: {e}")
            self.historial = []
    
    def guardar_historial(self):
        """Guarda todos los cálculos en el archivo JSON"""
        try:
            with open(self.archivo_historial, 'w', encoding='utf-8') as f:
                # Convierte cada objeto a diccionario antes de guardar
                json.dump([calc.to_dict() for calc in self.historial], f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"❌ Error al guardar el historial: {e}")
    
    def agregar_calculo(self, calculo: CalculoPropina):
        """Añade un nuevo cálculo al historial y lo guarda"""
        self.historial.append(calculo)
        self.guardar_historial()
    
    def obtener_estadisticas(self) -> Dict:
        """Calcula estadísticas básicas de todos los cálculos"""
        if not self.historial:
            return {"total_calculos": 0}
        
        # Extrae los valores necesarios para las estadísticas
        propinas = [calc.propina for calc in self.historial]
        montos = [calc.monto_original for calc in self.historial]
        
        return {
            "total_calculos": len(self.historial),
            "propina_promedio": sum(propinas) / len(propinas),
            "propina_maxima": max(propinas),
            "propina_minima": min(propinas),
            "monto_promedio": sum(montos) / len(montos),
            "total_propinas_acumuladas": sum(propinas),
            "total_gastado": sum(montos)
        }
    
    def exportar_excel(self, nombre_archivo: str = "historial_propinas.xlsx") -> bool:
        """
        Exporta el historial a un archivo EXCEL (XLSX) que se puede abrir con Excel.
        
        NOTA: Para usar esta función necesitas instalar openpyxl:
        pip install openpyxl
        
        Si no tienes openpyxl, se usará CSV como alternativa.
        """
        if not self.historial:
            print("📭 No hay datos para exportar")
            return False
        
        try:
            # Intenta usar openpyxl para crear un archivo Excel real
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Crear nuevo libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial de Propinas"
            
            # Definir estilos para el encabezado
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Definir bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Escribir encabezados (nombres de columnas)
            headers = ["Fecha", "Monto Original", "Moneda", "Propina %", 
                      "Calidad Servicio", "Propina", "Total", "Personas", "Por Persona"]
            
            for col, header in enumerate(headers, 1):
                celda = ws.cell(row=1, column=col, value=header)
                celda.font = header_font
                celda.fill = header_fill
                celda.alignment = header_alignment
                celda.border = thin_border
            
            # Escribir los datos de cada cálculo
            for fila, calculo in enumerate(self.historial, 2):
                # Convertir fecha a formato legible
                fecha_obj = datetime.fromisoformat(calculo.fecha)
                fecha_formateada = fecha_obj.strftime('%d/%m/%Y %H:%M:%S')
                
                # Escribir cada dato en su columna
                ws.cell(row=fila, column=1, value=fecha_formateada)
                ws.cell(row=fila, column=2, value=calculo.monto_original)
                ws.cell(row=fila, column=3, value=calculo.moneda)
                ws.cell(row=fila, column=4, value=calculo.porcentaje_propina)
                ws.cell(row=fila, column=5, value=calculo.calidad_servicio)
                ws.cell(row=fila, column=6, value=calculo.propina)
                ws.cell(row=fila, column=7, value=calculo.total)
                ws.cell(row=fila, column=8, value=calculo.num_personas)
                ws.cell(row=fila, column=9, value=calculo.total_por_persona)
                
                # Aplicar bordes y alineación a las celdas de datos
                for col in range(1, 10):
                    celda = ws.cell(row=fila, column=col)
                    celda.border = thin_border
                    if col in [2, 4, 6, 7, 9]:  # Columnas numéricas
                        celda.alignment = Alignment(horizontal="right")
                    else:
                        celda.alignment = Alignment(horizontal="left")
            
            # Ajustar automáticamente el ancho de las columnas
            for col in range(1, 10):
                col_letter = chr(64 + col)  # Convierte 1 a 'A', 2 a 'B', etc.
                ws.column_dimensions[col_letter].auto_size = True
            
            # Crear una segunda hoja con estadísticas
            ws_stats = wb.create_sheet("Estadísticas")
            
            # Obtener estadísticas
            stats = self.obtener_estadisticas()
            
            # Escribir estadísticas
            stats_data = [
                ["ESTADÍSTICAS DEL HISTORIAL", ""],
                ["", ""],
                ["Total de cálculos", stats.get("total_calculos", 0)],
                ["Monto promedio", f"{stats.get('monto_promedio', 0):.2f}"],
                ["Propina promedio", f"{stats.get('propina_promedio', 0):.2f}"],
                ["Propina más alta", f"{stats.get('propina_maxima', 0):.2f}"],
                ["Propina más baja", f"{stats.get('propina_minima', 0):.2f}"],
                ["Total en propinas", f"{stats.get('total_propinas_acumuladas', 0):.2f}"],
                ["Total gastado", f"{stats.get('total_gastado', 0):.2f}"]
            ]
            
            for fila, (desc, valor) in enumerate(stats_data, 1):
                ws_stats.cell(row=fila, column=1, value=desc)
                ws_stats.cell(row=fila, column=2, value=valor)
                
                # Dar formato a los títulos
                if "ESTADÍSTICAS" in desc:
                    ws_stats.cell(row=fila, column=1).font = Font(bold=True, size=14)
            
            # Ajustar ancho de columnas en hoja de estadísticas
            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 20
            
            # Guardar el archivo
            wb.save(nombre_archivo)
            print(f"✅ Archivo Excel guardado como: {nombre_archivo}")
            print(f"📂 Se crearon {len(self.historial)} registros en el archivo")
            return True
            
        except ImportError:
            # Si no está instalado openpyxl, usar CSV como alternativa
            print("⚠️ openpyxl no está instalado. Usando formato CSV...")
            print("💡 Para exportar a Excel, instala: pip install openpyxl")
            return self.exportar_csv(nombre_archivo.replace('.xlsx', '.csv'))
        except Exception as e:
            print(f"❌ Error al exportar: {e}")
            return False
    
    def exportar_csv(self, nombre_archivo: str = "historial_propinas.csv") -> bool:
        """Exporta el historial a CSV (compatible con Excel)"""
        try:
            with open(nombre_archivo, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # Escribir encabezados
                writer.writerow(["Fecha", "Monto Original", "Moneda", "Propina %", 
                               "Calidad Servicio", "Propina", "Total", "Personas", "Por Persona"])
                
                # Escribir datos
                for calculo in self.historial:
                    fecha_obj = datetime.fromisoformat(calculo.fecha)
                    fecha_formateada = fecha_obj.strftime('%d/%m/%Y %H:%M:%S')
                    
                    writer.writerow([
                        fecha_formateada,
                        calculo.monto_original,
                        calculo.moneda,
                        calculo.porcentaje_propina,
                        calculo.calidad_servicio,
                        calculo.propina,
                        calculo.total,
                        calculo.num_personas,
                        calculo.total_por_persona
                    ])
            
            print(f"✅ Archivo CSV guardado como: {nombre_archivo}")
            print(f"📂 Puedes abrir este archivo con Excel")
            return True
        except Exception as e:
            print(f"❌ Error al exportar CSV: {e}")
            return False


# =================================================================================
# CALCULADORA PRINCIPAL
# =================================================================================

class CalculadoraPropina:
    """Calculadora principal que realiza todos los cálculos"""
    
    def __init__(self):
        """Inicializa la calculadora con valores por defecto"""
        self.moneda_actual = Moneda.USD
        self.gestor_historial = GestorHistorial()
    
    def calcular(self, monto: float, calidad: CalidadServicio = None, 
                 porcentaje_personalizado: float = None, num_personas: int = 1) -> Optional[CalculoPropina]:
        """
        Realiza el cálculo de propina
        
        Args:
            monto: Monto de la cuenta
            calidad: Calidad del servicio (opciones predefinidas)
            porcentaje_personalizado: Porcentaje manual (sobrescribe calidad)
            num_personas: Número de personas para dividir (1 si no divide)
        """
        # Validar monto
        if monto <= 0:
            print("❌ El monto debe ser mayor a cero")
            return None
        
        # Validar número de personas
        if num_personas < 1:
            print("❌ El número de personas debe ser al menos 1")
            return None
        
        # Determinar el porcentaje a usar
        if porcentaje_personalizado is not None:
            # Usar porcentaje personalizado
            if porcentaje_personalizado < 0 or porcentaje_personalizado > 100:
                print("❌ El porcentaje debe estar entre 0 y 100")
                return None
            porcentaje = porcentaje_personalizado
            calidad_desc = f"Personalizado ({porcentaje}%)"
        elif calidad:
            # Usar calidad predefinida
            porcentaje = calidad.porcentaje
            calidad_desc = calidad.descripcion
        else:
            # Por defecto 15%
            porcentaje = 15
            calidad_desc = "Estándar (15%)"
        
        # Realizar cálculos
        propina = monto * (porcentaje / 100)
        total = monto + propina
        total_por_persona = total / num_personas
        
        # Redondear a 2 decimales
        propina = round(propina, 2)
        total = round(total, 2)
        total_por_persona = round(total_por_persona, 2)
        
        # Crear objeto con el resultado
        resultado = CalculoPropina(
            fecha=datetime.now().isoformat(),
            monto_original=round(monto, 2),
            moneda=self.moneda_actual.codigo,
            porcentaje_propina=porcentaje,
            calidad_servicio=calidad_desc,
            propina=propina,
            total=total,
            num_personas=num_personas,
            total_por_persona=total_por_persona
        )
        
        # Guardar en historial
        self.gestor_historial.agregar_calculo(resultado)
        
        return resultado
    
    def cambiar_moneda(self, nueva_moneda: Moneda):
        """Cambia la moneda actual"""
        self.moneda_actual = nueva_moneda
        print(f"✅ Moneda cambiada a {nueva_moneda.codigo} {nueva_moneda.simbolo}")
    
    def mostrar_resultado(self, resultado: CalculoPropina):
        """Muestra el resultado de forma clara y legible"""
        simbolo = self.moneda_actual.simbolo
        
        print("\n" + "=" * 55)
        print("                 📊 RESULTADO")
        print("=" * 55)
        print(f"💰 Monto original:     {simbolo}{resultado.monto_original}")
        print(f"⭐ Calidad servicio:   {resultado.calidad_servicio}")
        print(f"💵 Propina:           {simbolo}{resultado.propina}")
        print(f"🌟 Total a pagar:      {simbolo}{resultado.total}")
        
        if resultado.num_personas > 1:
            print(f"👥 Entre {resultado.num_personas} personas:")
            print(f"👤 Cada persona paga: {simbolo}{resultado.total_por_persona}")
        
        print("=" * 55)


# =================================================================================
# INTERFAZ DE USUARIO
# =================================================================================

class InterfazUsuario:
    """Maneja la interacción con el usuario en consola"""
    
    def __init__(self):
        """Inicializa la interfaz y la calculadora"""
        self.calculadora = CalculadoraPropina()
    
    def limpiar_pantalla(self):
        """Limpia la pantalla de la consola"""
        os.system('cls' if os.name == 'nt' else 'clear')
    
    def mostrar_menu(self):
        """Muestra el menú principal"""
        print("\n" + "=" * 55)
        print("        🍽️  CALCULADORA DE PROPINAS  🍽️")
        print("=" * 55)
        print(f"💰 Moneda actual: {self.calculadora.moneda_actual.codigo} {self.calculadora.moneda_actual.simbolo}")
        print("-" * 55)
        print("1. 📝 Calcular propina")
        print("2. 📜 Ver historial")
        print("3. 📊 Ver estadísticas")
        print("4. 💱 Cambiar moneda")
        print("5. 📤 Exportar a Excel/CSV")
        print("6. ❌ Salir")
        print("=" * 55)
    
    def calcular_propina(self):
        """Flujo para calcular una nueva propina"""
        self.limpiar_pantalla()
        print("\n" + "=" * 55)
        print("              📝 NUEVO CÁLCULO")
        print("=" * 55)
        
        try:
            # Pedir monto
            simbolo = self.calculadora.moneda_actual.simbolo
            monto = float(input(f"💰 Monto de la cuenta ({simbolo}): "))
            
            # Mostrar opciones de calidad
            print("\n⭐ Calidad del servicio:")
            opciones = list(CalidadServicio)
            for i, calidad in enumerate(opciones, 1):
                print(f"   {i}. {calidad.descripcion}")
            print(f"   {len(opciones) + 1}. Personalizado")
            
            opcion = int(input("\nSelecciona una opción: "))
            
            calidad = None
            porcentaje_personalizado = None
            
            if 1 <= opcion <= len(opciones):
                calidad = opciones[opcion - 1]
            elif opcion == len(opciones) + 1:
                porcentaje_personalizado = float(input("Porcentaje personalizado (%): "))
            else:
                print("❌ Opción no válida")
                input("Presiona Enter...")
                return
            
            # Preguntar si divide cuenta
            dividir = input("\n👥 ¿Dividir la cuenta? (s/n): ").lower()
            num_personas = 1
            if dividir == 's':
                num_personas = int(input("Número de personas: "))
            
            # Realizar cálculo
            resultado = self.calculadora.calcular(
                monto=monto,
                calidad=calidad,
                porcentaje_personalizado=porcentaje_personalizado,
                num_personas=num_personas
            )
            
            if resultado:
                self.calculadora.mostrar_resultado(resultado)
            
        except ValueError:
            print("❌ Error: Ingresa números válidos")
        except Exception as e:
            print(f"❌ Error: {e}")
        
        input("\nPresiona Enter para continuar...")
    
    def ver_historial(self):
        """Muestra el historial de cálculos"""
        self.limpiar_pantalla()
        print("\n" + "=" * 55)
        print("              📜 HISTORIAL DE CÁLCULOS")
        print("=" * 55)
        
        historial = self.calculadora.gestor_historial.historial
        
        if not historial:
            print("📭 No hay cálculos guardados aún")
        else:
            # Mostrar últimos 10 cálculos
            for i, calc in enumerate(historial[-10:], 1):
                fecha = datetime.fromisoformat(calc.fecha).strftime('%d/%m/%Y %H:%M')
                print(f"\n{i}. 📅 {fecha}")
                print(f"   💰 {calc.monto_original} {calc.moneda} | {calc.calidad_servicio}")
                print(f"   💵 Propina: {calc.propina} {calc.moneda} | Total: {calc.total} {calc.moneda}")
                if calc.num_personas > 1:
                    print(f"   👥 {calc.num_personas} personas: {calc.total_por_persona} {calc.moneda} c/u")
        
        print("\n" + "=" * 55)
        input("Presiona Enter para continuar...")
    
    def ver_estadisticas(self):
        """Muestra estadísticas del historial"""
        self.limpiar_pantalla()
        print("\n" + "=" * 55)
        print("              📊 ESTADÍSTICAS")
        print("=" * 55)
        
        stats = self.calculadora.gestor_historial.obtener_estadisticas()
        simbolo = self.calculadora.moneda_actual.simbolo
        
        if stats["total_calculos"] == 0:
            print("📭 No hay suficientes datos para estadísticas")
        else:
            print(f"📊 Total de cálculos: {stats['total_calculos']}")
            print(f"💰 Monto promedio: {simbolo}{stats['monto_promedio']:.2f}")
            print(f"💵 Propina promedio: {simbolo}{stats['propina_promedio']:.2f}")
            print(f"📈 Propina más alta: {simbolo}{stats['propina_maxima']:.2f}")
            print(f"📉 Propina más baja: {simbolo}{stats['propina_minima']:.2f}")
            print(f"💰 Total en propinas: {simbolo}{stats['total_propinas_acumuladas']:.2f}")
            print(f"💳 Total gastado: {simbolo}{stats['total_gastado']:.2f}")
        
        print("=" * 55)
        input("Presiona Enter para continuar...")
    
    def cambiar_moneda(self):
        """Menú para cambiar la moneda"""
        self.limpiar_pantalla()
        print("\n" + "=" * 55)
        print("              💱 CAMBIAR MONEDA")
        print("=" * 55)
        
        monedas = list(Moneda)
        for i, moneda in enumerate(monedas, 1):
            print(f"{i}. {moneda.codigo} ({moneda.simbolo})")
        
        try:
            opcion = int(input("\nSelecciona una moneda: "))
            if 1 <= opcion <= len(monedas):
                self.calculadora.cambiar_moneda(monedas[opcion - 1])
            else:
                print("❌ Opción no válida")
        except ValueError:
            print("❌ Ingresa un número válido")
        
        input("\nPresiona Enter para continuar...")
    
    def exportar_datos(self):
        """Exporta el historial a Excel o CSV"""
        self.limpiar_pantalla()
        print("\n" + "=" * 55)
        print("              📤 EXPORTAR DATOS")
        print("=" * 55)
        print("1. Exportar a Excel (.xlsx) - Recomendado")
        print("2. Exportar a CSV (.csv) - Compatible con Excel")
        print("3. Volver")
        
        opcion = input("\nSelecciona una opción: ")
        
        if opcion == '1':
            nombre = input("Nombre del archivo [historial_propinas]: ").strip()
            if not nombre:
                nombre = "historial_propinas"
            nombre += ".xlsx"
            self.calculadora.gestor_historial.exportar_excel(nombre)
        elif opcion == '2':
            nombre = input("Nombre del archivo [historial_propinas]: ").strip()
            if not nombre:
                nombre = "historial_propinas"
            nombre += ".csv"
            self.calculadora.gestor_historial.exportar_csv(nombre)
        elif opcion == '3':
            return
        
        input("\nPresiona Enter para continuar...")
    
    def ejecutar(self):
        """Bucle principal del programa"""
        while True:
            self.limpiar_pantalla()
            self.mostrar_menu()
            
            opcion = input("\n🔧 Opción: ")
            
            if opcion == '1':
                self.calcular_propina()
            elif opcion == '2':
                self.ver_historial()
            elif opcion == '3':
                self.ver_estadisticas()
            elif opcion == '4':
                self.cambiar_moneda()
            elif opcion == '5':
                self.exportar_datos()
            elif opcion == '6':
                print("\n👋 ¡Gracias por usar la calculadora!")
                break
            else:
                print("❌ Opción no válida")
                input("Presiona Enter...")


# =================================================================================
# PUNTO DE ENTRADA DEL PROGRAMA
# =================================================================================

if __name__ == "__main__":
    # Crear y ejecutar la interfaz
    app = InterfazUsuario()
    app.ejecutar()
